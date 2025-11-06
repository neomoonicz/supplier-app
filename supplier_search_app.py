import streamlit as st
import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

# -----------------------------------------------------
# 🧩 Streamlit Page Setup
# -----------------------------------------------------
st.set_page_config(page_title="Supplier Search & Summary", layout="wide")
st.title("📊 Supplier Search and Summary App")

# -----------------------------------------------------
# ⚙️ Cached Excel Loader for Performance
# -----------------------------------------------------
@st.cache_data
def load_excel(file):
    """Read Excel once and cache it for faster live search."""
    return pd.read_excel(file, header=None)

# -----------------------------------------------------
# 📂 File Upload
# -----------------------------------------------------
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        df = load_excel(uploaded_file)

        # --- Automatic header row detection ---
        header_row_index = None
        for i in range(len(df)):
            row_values = [str(x).strip().upper() for x in df.iloc[i].tolist()]
            if "NAME OF SUPPLIERS" in row_values and "TOTAL AMOUNT PAID" in row_values:
                header_row_index = i
                break

        if header_row_index is None:
            st.error("Could not find the header row (with 'NAME OF SUPPLIERS' and 'TOTAL AMOUNT PAID').")
        else:
            # Set header and clean
            df.columns = df.iloc[header_row_index]
            df = df[header_row_index + 1:].reset_index(drop=True)
            df.columns = [str(c).strip().upper() for c in df.columns]

            # --- Auto column detection ---
            supplier_col = next((c for c in df.columns if c == "NAME OF SUPPLIERS"), None)
            tin_col = next((c for c in df.columns if c == "TIN"), None)
            total_col = next((c for c in df.columns if c == "TOTAL AMOUNT PAID"), None)

            # --- Manual fallback if detection fails ---
            if not all([supplier_col, total_col]):
                st.warning("Automatic detection failed. Please select columns manually:")
                supplier_col = st.selectbox("Supplier Name column:", df.columns)
                tin_col = st.selectbox("TIN column:", ["N/A"] + df.columns.tolist())
                total_col = st.selectbox("Total Amount Paid column:", df.columns)

            # --- Data Cleaning ---
            df = df[[supplier_col, tin_col, total_col]].copy()
            df.columns = ["Supplier Name", "TIN", "Total Amount Paid"]

            # Remove rows where all three are empty
            df = df.dropna(how='all', subset=["Supplier Name", "TIN", "Total Amount Paid"])
            df = df[~((df["Supplier Name"].astype(str).str.strip() == "") &
                      (df["TIN"].astype(str).str.strip() == "") &
                      (df["Total Amount Paid"].astype(str).str.strip() == ""))]

            # --- Normalize TIN for searching ---
            def clean_tin(tin):
                if pd.isna(tin):
                    return ""
                return re.sub(r"[^A-Za-z0-9]", "", str(tin)).lower()

            df["TIN_CLEAN"] = df["TIN"].apply(clean_tin)

            # -----------------------------------------------------
            # 🔍 Live Search Input
            # -----------------------------------------------------
            search_query = st.text_input(
                "Search Supplier Name or TIN:",
                placeholder="Type supplier name or TIN to filter..."
            )

            if search_query:
                cleaned_query = search_query.strip().lower()
                filtered_df = df[
                    df["Supplier Name"].astype(str).str.lower().str.contains(cleaned_query, na=False) |
                    df["TIN_CLEAN"].astype(str).str.lower().str.contains(cleaned_query, na=False)
                ]
            else:
                filtered_df = df

            # -----------------------------------------------------
            # 📊 Display Filtered Results
            # -----------------------------------------------------
            if filtered_df.empty:
                st.warning("No matching records found.")
            else:
                # Add TOTAL row
                total_sum = pd.to_numeric(filtered_df["Total Amount Paid"], errors='coerce').sum()
                total_row = pd.DataFrame({
                    "Supplier Name": ["TOTAL"],
                    "TIN": [""],
                    "Total Amount Paid": [total_sum]
                })
                display_df = pd.concat(
                    [filtered_df[["Supplier Name", "TIN", "Total Amount Paid"]], total_row],
                    ignore_index=True
                )

                # Style total row in Streamlit table
                def highlight_total_row(row):
                    if str(row["Supplier Name"]).strip().upper() == "TOTAL":
                        return ["font-weight: bold; font-size: 1.1em; background-color: #f2f2f2;" for _ in row]
                    return [""] * len(row)

                styled_df = display_df.style.format(
                    {"Total Amount Paid": "{:,.2f}"}
                ).apply(highlight_total_row, axis=1)

                st.subheader("Search Results Summary")
                st.dataframe(styled_df, use_container_width=True, hide_index=True)
                st.markdown("---")
                st.write(f"**Number of Entries (excluding total):** {len(filtered_df)}")

                # -----------------------------------------------------
                # 📤 Excel Export with Highlighted TOTAL
                # -----------------------------------------------------
                def export_to_excel(df):
                    output = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Supplier Summary"

                    # Write DataFrame rows
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)

                    # --- Header styling ---
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # --- Highlight TOTAL row (gray + bold) ---
                    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        supplier_name = str(row[0].value).strip().upper()
                        if supplier_name == "TOTAL":
                            for cell in row:
                                cell.font = Font(bold=True)
                                cell.fill = gray_fill

                    # --- Auto column width ---
                    for i, col in enumerate(ws.columns, start=1):
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                cell_value = str(cell.value)
                                if len(cell_value) > max_length:
                                    max_length = len(cell_value)
                            except Exception:
                                pass
                        ws.column_dimensions[column].width = max_length + (4 if i == 1 else 2)

                    wb.save(output)
                    output.seek(0)
                    return output

                excel_file = export_to_excel(display_df)

                st.download_button(
                    label="📥 Export to Excel",
                    data=excel_file,
                    file_name="Supplier_Summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"Error processing file: {e}")

else:
    st.info("Please upload an Excel file to start.")
